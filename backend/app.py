import os
import json
from datetime import datetime
from flask import Flask, request, jsonify, send_file
import io

# Importar lógica local
from .features import parsear_fasta, parsear_csv, limpiar_y_validar_secuencia, extraer_todas_las_caracteristicas
from .model_manager import ModelManager

# Definir rutas base - Compatible con local y Render
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")

# Crear directorios si no existen
os.makedirs(os.path.join(BASE_DIR, "backend", "data"), exist_ok=True)

HISTORY_PATH = os.path.join(BASE_DIR, "backend", "data", "history.json")

# Configurar Flask con archivos estáticos
app = Flask(__name__, static_folder=FRONTEND_DIR, static_url_path="")
# Aumentar timeout para predicciones complejas
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# Inicializar gestor de modelos al arrancar el servidor (con recarga dinámica de métricas)
model_manager = ModelManager()
try:
    model_manager.inicializar()
    print("✓ Modelos inicializados correctamente")
except Exception as e:
    print(f"✗ ERROR CRÍTICO al inicializar modelos: {e}")
    import traceback
    traceback.print_exc()
    raise

def guardar_en_historial(registro):
    historial = []
    if os.path.exists(HISTORY_PATH):
        try:
            with open(HISTORY_PATH, "r", encoding="utf-8") as f:
                historial = json.load(f)
        except Exception:
            historial = []
            
    historial.insert(0, registro)  # Insertar al inicio para que aparezca primero
    # Limitar el historial a las últimas 100 muestras para no sobrecargar el archivo
    historial = historial[:100]
    
    with open(HISTORY_PATH, "w", encoding="utf-8") as f:
        json.dump(historial, f, indent=2, ensure_ascii=False)

@app.route("/")
def index():
    return app.send_static_file("index.html")

@app.route("/api/predict", methods=["POST"])
def predict():
    data = request.get_json()
    if not data or "secuencia" not in data:
        return jsonify({"success": False, "error": "No se recibió ninguna secuencia de ADN o archivo."}), 400
        
    contenido = data["secuencia"].strip()
    
    # Detectar si el contenido tiene estructura de CSV
    es_csv = False
    if not contenido.startswith('>'):
        lineas = contenido.split('\n')
        if len(lineas) > 1:
            header = lineas[0]
            if ',' in header or ';' in header:
                es_csv = True
                
    if es_csv:
        try:
            import csv
            # Primero, detectar si es un CSV de características precalculadas
            f_check = io.StringIO(contenido)
            # Leer el header para ver qué columnas contiene
            lineas = contenido.split('\n')
            header_line = lineas[0].lower()
            delimitador = ';' if ';' in header_line else ','
            
            reader_check = csv.DictReader(f_check, delimiter=delimitador)
            fieldnames_lower = [col.lower().strip() for col in (reader_check.fieldnames or [])]
            
            # Columnas clave que indican que contiene características precalculadas
            caracteristicas_clave = ["freq_a", "freq_t", "freq_c", "freq_g", "contenido_gc"]
            es_csv_caracteristicas = all(col in fieldnames_lower for col in caracteristicas_clave)
            
            if es_csv_caracteristicas:
                # Buscar columna de ID
                col_id = None
                nombres_id = ["id", "nombre", "name", "identifier", "muestra", "sample"]
                for col in reader_check.fieldnames or []:
                    if col.lower().strip() in nombres_id:
                        col_id = col
                        break
                
                # Mapear los nombres de columnas exactos de la CSV a los esperados por el modelo (case-sensitive)
                col_mapping = {}
                for col in reader_check.fieldnames or []:
                    col_strip = col.strip()
                    col_lower = col_strip.lower()
                    col_mapping[col_lower] = col_strip
                
                resultados_batch = []
                errores = []
                
                for i, row in enumerate(reader_check):
                    id_muestra = str(row.get(col_id, f"muestra_csv_{i+1}")).strip() if col_id else f"muestra_csv_{i+1}"
                    if not id_muestra:
                        id_muestra = f"muestra_csv_{i+1}"
                        
                    try:
                        features = {}
                        # Mapear las 69 características numéricas
                        for feat_name in model_manager.feature_names:
                            feat_name_lower = feat_name.lower()
                            actual_col = col_mapping.get(feat_name_lower)
                            if actual_col and row.get(actual_col) is not None:
                                features[feat_name] = float(row.get(actual_col))
                            else:
                                raise ValueError(f"Falta la columna de característica obligatoria: '{feat_name}'")
                        
                        # Realizar predicción
                        predicciones = model_manager.predecir(features)
                        
                        fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        resultado_rf = predicciones["random_forest"]["resultado"]
                        prob_rf = predicciones["random_forest"]["probabilidad"]
                        
                        registro_historial = {
                            "id": id_muestra,
                            "fecha": fecha_hora,
                            "longitud": 0,  # No hay secuencia de nucleótidos cruda
                            "gc_content": round(features["Contenido_GC"], 2),
                            "resultado": resultado_rf,
                            "probabilidad": round(prob_rf * 100, 2),
                            "modelo": "Random Forest",
                            "detalles": {
                                "logistic_regression": {
                                    "resultado": predicciones["logistic_regression"]["resultado"],
                                    "probabilidad": round(predicciones["logistic_regression"]["probabilidad"] * 100, 2)
                                },
                                "random_forest": {
                                    "resultado": resultado_rf,
                                    "probabilidad": round(prob_rf * 100, 2)
                                }
                            }
                        }
                        
                        guardar_en_historial(registro_historial)
                        resultados_batch.append(registro_historial)
                        
                    except Exception as row_err:
                        errores.append(f"Muestra '{id_muestra}': {str(row_err)}")
                        
                if not resultados_batch:
                    error_detallado = "No se pudo procesar ninguna muestra válida con características. Errores: " + " | ".join(errores)
                    return jsonify({"success": False, "error": error_detallado}), 400
                    
                return jsonify({
                    "success": True,
                    "batch": True,
                    "data": resultados_batch,
                    "errors": errores
                })
                
            else:
                muestras = parsear_csv(contenido)
                if not muestras:
                    return jsonify({
                        "success": False, 
                        "error": "El archivo CSV no contiene columnas legibles de secuencias de ADN (se buscan cabeceras como 'Cadena_ADN', 'Sequence', 'ADN', 'seq', etc.)."
                    }), 400
                    
                resultados_batch = []
                errores = []
                
                for id_muestra, secuencia_cruda in muestras:
                    es_valido, mensaje_error, secuencia_limpia = limpiar_y_validar_secuencia(secuencia_cruda)
                    if not es_valido:
                        errores.append(f"Muestra '{id_muestra}': {mensaje_error}")
                        continue
                    
                    # Extraer características
                    features = extraer_todas_las_caracteristicas(secuencia_limpia)
                    
                    # Realizar predicciones
                    predicciones = model_manager.predecir(features)
                    
                    fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    resultado_rf = predicciones["random_forest"]["resultado"]
                    prob_rf = predicciones["random_forest"]["probabilidad"]
                    
                    registro_historial = {
                        "id": id_muestra,
                        "fecha": fecha_hora,
                        "longitud": len(secuencia_limpia),
                        "gc_content": round(features["Contenido_GC"], 2),
                        "resultado": resultado_rf,
                        "probabilidad": round(prob_rf * 100, 2),
                        "modelo": "Random Forest",
                        "detalles": {
                            "logistic_regression": {
                                "resultado": predicciones["logistic_regression"]["resultado"],
                                "probabilidad": round(predicciones["logistic_regression"]["probabilidad"] * 100, 2)
                            },
                            "random_forest": {
                                "resultado": resultado_rf,
                                "probabilidad": round(prob_rf * 100, 2)
                            }
                        }
                    }
                    
                    guardar_en_historial(registro_historial)
                    resultados_batch.append(registro_historial)
                    
                if not resultados_batch:
                    error_detallado = "No se pudo procesar ninguna secuencia válida del CSV. Errores: " + " | ".join(errores)
                    return jsonify({"success": False, "error": error_detallado}), 400
                    
                return jsonify({
                    "success": True,
                    "batch": True,
                    "data": resultados_batch,
                    "errors": errores
                })
            
        except Exception as e:
            return jsonify({"success": False, "error": f"Error procesando el lote CSV: {str(e)}"}), 500
            
    else:
        # Procesamiento normal de secuencia única (FASTA o texto plano)
        id_muestra, secuencia_cruda = parsear_fasta(contenido)
        
        # Validar formato
        es_valido, mensaje_error, secuencia_limpia = limpiar_y_validar_secuencia(secuencia_cruda)
        if not es_valido:
            return jsonify({"success": False, "error": mensaje_error}), 400
            
        try:
            # Extraer características
            features = extraer_todas_las_caracteristicas(secuencia_limpia)
            
            # Realizar predicciones
            predicciones = model_manager.predecir(features)
            
            # Estructurar resultado
            fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            resultado_rf = predicciones["random_forest"]["resultado"]
            prob_rf = predicciones["random_forest"]["probabilidad"]
            
            registro_historial = {
                "id": id_muestra,
                "fecha": fecha_hora,
                "longitud": len(secuencia_limpia),
                "gc_content": round(features["Contenido_GC"], 2),
                "resultado": resultado_rf,
                "probabilidad": round(prob_rf * 100, 2),
                "modelo": "Random Forest",
                "detalles": {
                    "logistic_regression": {
                        "resultado": predicciones["logistic_regression"]["resultado"],
                        "probabilidad": round(predicciones["logistic_regression"]["probabilidad"] * 100, 2)
                    },
                    "random_forest": {
                        "resultado": resultado_rf,
                        "probabilidad": round(prob_rf * 100, 2)
                    }
                }
            }
            
            # Guardar en el historial
            guardar_en_historial(registro_historial)
            
            return jsonify({
                "success": True,
                "batch": False,
                "data": registro_historial
            })
            
        except Exception as e:
            return jsonify({"success": False, "error": f"Error interno en el procesamiento: {str(e)}"}), 500

@app.route("/api/metrics", methods=["GET"])
def get_metrics():
    if not model_manager.metrics:
        return jsonify({"success": False, "error": "Las métricas no han sido cargadas."}), 500
    return jsonify({"success": True, "data": model_manager.metrics})

@app.route("/api/history", methods=["GET"])
def get_history():
    historial = []
    if os.path.exists(HISTORY_PATH):
        try:
            with open(HISTORY_PATH, "r", encoding="utf-8") as f:
                historial = json.load(f)
        except Exception:
            historial = []
    return jsonify({"success": True, "data": historial})

@app.route("/api/export", methods=["POST"])
def export_report():
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "error": "Datos inválidos para exportar."}), 400
        
    formato = data.get("format", "csv").lower()
    registro = data.get("record")
    
    if not registro:
        return jsonify({"success": False, "error": "No se encontraron datos del registro de análisis."}), 400

    if formato == "csv":
        # Generar CSV plano
        output = io.StringIO()
        output.write("REPORTE DE DIAGNÓSTICO BACTERIANO - RESISTENCIA A PENICILINA\n")
        output.write(f"Fecha del Reporte: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        output.write("DATOS DE LA MUESTRA\n")
        output.write(f"ID Muestra;{registro.get('id')}\n")
        output.write(f"Fecha Analisis;{registro.get('fecha')}\n")
        output.write(f"Longitud de Secuencia (bp);{registro.get('longitud')}\n")
        output.write(f"Contenido GC (%);{registro.get('gc_content')}\n\n")
        
        output.write("RESULTADOS DEL DIAGNÓSTICO (MODELO RECOMENDADO)\n")
        output.write(f"Modelo Utilizado;{registro.get('modelo')}\n")
        output.write(f"Diagnóstico Final;{registro.get('resultado')}\n")
        output.write(f"Confianza de Predicción;{registro.get('probabilidad')}%\n\n")
        
        output.write("DETALLE DE MODELOS COMPARATIVOS\n")
        output.write("Modelo;Prediccion;Confianza\n")
        
        detalles = registro.get("detalles", {})
        rf_det = detalles.get("random_forest", {})
        lr_det = detalles.get("logistic_regression", {})
        
        output.write(f"Random Forest;{rf_det.get('resultado')};{rf_det.get('probabilidad')}%\n")
        output.write(f"Regresión Logística;{lr_det.get('resultado')};{lr_det.get('probabilidad')}%\n")
        
        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8-sig')) # BOM para Excel en español
        mem.seek(0)
        
        filename = f"Reporte_{registro.get('id')}.csv"
        return send_file(
            mem,
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename
        )
        
    elif formato == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte Diagnóstico"
            
            # Setup styles
            title_font = Font(name="Arial", size=14, bold=True, color="0F766E")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            bold_font = Font(name="Arial", size=10, bold=True)
            regular_font = Font(name="Arial", size=10)
            
            header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )
            
            # Title
            ws["A1"] = "REPORTE DE DIAGNÓSTICO BACTERIANO"
            ws["A1"].font = title_font
            ws["A2"] = "Detección de resistencia a la Penicilina en E. coli"
            ws["A2"].font = Font(name="Arial", size=11, italic=True)
            ws["A3"] = f"Fecha del Reporte: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws["A3"].font = regular_font
            
            # Section: Sample
            ws["A5"] = "DATOS DE LA MUESTRA"
            ws["A5"].font = bold_font
            
            sample_data = [
                ("ID Muestra", registro.get('id')),
                ("Fecha Analisis", registro.get('fecha')),
                ("Longitud de Secuencia (bp)", registro.get('longitud') if registro.get('longitud') else "Precalculada"),
                ("Contenido GC (%)", registro.get('gc_content'))
            ]
            for row_idx, (k, v) in enumerate(sample_data, start=6):
                ws.cell(row=row_idx, column=1, value=k).font = bold_font
                ws.cell(row=row_idx, column=2, value=v).font = regular_font
                ws.cell(row=row_idx, column=1).border = thin_border
                ws.cell(row=row_idx, column=2).border = thin_border
                
            # Section: Results
            ws["A11"] = "RESULTADOS DEL DIAGNÓSTICO (MODELO RECOMENDADO)"
            ws["A11"].font = bold_font
            
            results_data = [
                ("Modelo Utilizado", registro.get('modelo')),
                ("Diagnóstico Final", registro.get('resultado')),
                ("Confianza de Predicción", f"{registro.get('probabilidad')}%")
            ]
            for row_idx, (k, v) in enumerate(results_data, start=12):
                ws.cell(row=row_idx, column=1, value=k).font = bold_font
                c = ws.cell(row=row_idx, column=2, value=v)
                c.font = bold_font
                if k == "Diagnóstico Final":
                    c.font = Font(name="Arial", size=10, bold=True, color="B91C1C" if v == "Resistente" else "15803D")
                ws.cell(row=row_idx, column=1).border = thin_border
                ws.cell(row=row_idx, column=2).border = thin_border
                ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

            # Section: Detailed models
            ws["A16"] = "DETALLE DE MODELOS COMPARATIVOS"
            ws["A16"].font = bold_font
            
            headers = ["Modelo", "Prediccion", "Confianza"]
            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(row=17, column=col_idx, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center")
                c.border = thin_border
                
            detalles = registro.get("detalles", {})
            rf_det = detalles.get("random_forest", {})
            lr_det = detalles.get("logistic_regression", {})
            
            comparison_rows = [
                ("Random Forest (Primario)", rf_det.get('resultado'), f"{rf_det.get('probabilidad')}%"),
                ("Regresión Logística (Secundario)", lr_det.get('resultado'), f"{lr_det.get('probabilidad')}%")
            ]
            for row_idx, r in enumerate(comparison_rows, start=18):
                for col_idx, val in enumerate(r, start=1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.font = regular_font
                    c.border = thin_border
                    c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")

            # Column auto-fit
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            mem = io.BytesIO()
            wb.save(mem)
            mem.seek(0)
            
            filename = f"Reporte_{registro.get('id')}.xlsx"
            return send_file(
                mem,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename
            )
        except Exception as e:
            return jsonify({"success": False, "error": f"Error al generar Excel nativo: {str(e)}"}), 500
        
    elif formato == "pdf":
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            pdf_buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                pdf_buffer, 
                pagesize=letter,
                rightMargin=54, leftMargin=54, topMargin=54, bottomMargin=54
            )
            
            styles = getSampleStyleSheet()
            
            # Estilos personalizados
            title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontSize=20,
                textColor=colors.HexColor('#0F766E'), # Teal hospitalario
                spaceAfter=15,
                alignment=1 # Centrado
            )
            
            h2_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#1E293B'),
                spaceBefore=12,
                spaceAfter=6,
                borderPadding=4
            )
            
            body_style = ParagraphStyle(
                'BodyTextCustom',
                parent=styles['Normal'],
                fontSize=10.5,
                leading=14,
                textColor=colors.HexColor('#334155')
            )
            
            story = []
            
            # Encabezado
            story.append(Paragraph("REPORTE CLÍNICO DE DIAGNÓSTICO BACTERIANO", title_style))
            story.append(Paragraph("Predicción de resistencia a la Penicilina en cepas de *Escherichia coli*", styles['Normal']))
            story.append(Spacer(1, 15))
            
            # Información de la muestra
            story.append(Paragraph("Información de la Muestra", h2_style))
            datos_muestra = [
                [Paragraph("<b>ID de Muestra:</b>", body_style), Paragraph(str(registro.get('id')), body_style)],
                [Paragraph("<b>Fecha de Análisis:</b>", body_style), Paragraph(str(registro.get('fecha')), body_style)],
                [Paragraph("<b>Longitud de Secuencia (bp):</b>", body_style), Paragraph(f"{registro.get('longitud')} pb" if registro.get('longitud') else "Precalculada", body_style)],
                [Paragraph("<b>Contenido GC (%):</b>", body_style), Paragraph(f"{registro.get('gc_content')}%", body_style)]
            ]
            t_muestra = Table(datos_muestra, colWidths=[180, 320])
            t_muestra.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ]))
            story.append(t_muestra)
            story.append(Spacer(1, 15))
            
            # Resultado Diagnóstico
            story.append(Paragraph("Resultados del Análisis (Modelo de Consenso)", h2_style))
            color_res = '#B91C1C' if registro.get('resultado') == 'Resistente' else '#15803D'
            res_style = ParagraphStyle(
                'ResultVal',
                parent=body_style,
                fontName='Helvetica-Bold',
                fontSize=12,
                textColor=colors.HexColor(color_res)
            )
            
            datos_resultado = [
                [Paragraph("<b>Diagnóstico Final:</b>", body_style), Paragraph(str(registro.get('resultado')), res_style)],
                [Paragraph("<b>Nivel de Confianza:</b>", body_style), Paragraph(f"{registro.get('probabilidad')}% ({registro.get('resultado')})", body_style)],
                [Paragraph("<b>Modelo Determinante:</b>", body_style), Paragraph(str(registro.get('modelo')), body_style)]
            ]
            t_res = Table(datos_resultado, colWidths=[180, 320])
            t_res.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F1F5F9')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 8),
                ('TOPPADDING', (0,0), (-1,-1), 8),
                ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ]))
            story.append(t_res)
            story.append(Spacer(1, 15))
            
            # Detalle comparativo de modelos
            story.append(Paragraph("Comparación de Modelos Predictivos", h2_style))
            detalles = registro.get("detalles", {})
            rf_det = detalles.get("random_forest", {})
            lr_det = detalles.get("logistic_regression", {})
            
            datos_tabla_modelos = [
                ["Modelo de Machine Learning", "Predicción", "Nivel de Confianza"],
                ["Random Forest (Primario)", str(rf_det.get('resultado')), f"{rf_det.get('probabilidad')}%"],
                ["Regresión Logística (Secundario)", str(lr_det.get('resultado')), f"{lr_det.get('probabilidad')}%"]
            ]
            t_modelos = Table(datos_tabla_modelos, colWidths=[200, 150, 150])
            t_modelos.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F766E')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 10),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ]))
            story.append(t_modelos)
            
            story.append(Spacer(1, 25))
            story.append(Paragraph("<i>Nota: Este es un reporte bioinformático automatizado basado en modelos entrenados con secuencias genómicas experimentales de NCBI. Debe correlacionarse clínicamente en laboratorios de microbiología.</i>", styles['Italic']))
            
            doc.build(story)
            pdf_buffer.seek(0)
            
            return send_file(
                pdf_buffer,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=f"Reporte_{registro.get('id')}.pdf"
            )
            
        except ImportError:
            return jsonify({"success": False, "error": "La librería 'reportlab' no está instalada en el sistema para generar PDFs."}), 500
        except Exception as e:
            return jsonify({"success": False, "error": f"Error al generar PDF: {str(e)}"}), 500
            
    else:
        return jsonify({"success": False, "error": "Formato de exportación no soportado."}), 400

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)
